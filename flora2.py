#  会社概要をエクセルファイルに書き込む処理
#  新規作成  2024/3/3
#  Create by  乃木坂好きのITエンジニア

# ライブラリのインポート
import re
import openpyxl as op
import shutil
import os
import sys

#項目を取得
pattern1 = '<td class="title">'

#内容を取得
patturn2 = '<td class="content">'
patturn3 = '</td>'
patturn4 = '<br/>'

# pタグを取得
patturn5 = '<p>'

# データ切り替えレコード変数の定義
record1 = '<td class="content">'
record2 = '<td class="title">事業内容</td>'

# 入力ファイル
input_file =  "flora.txt"
# テンプレートエクセルファイル
template_excel_file = "会社概要.xlsx"
# 出力ファイル
output_file = "会社概要出力結果.xlsx"
# 会社名、ホールディング名判定カウント
w_count = 0
# pタグ会社名、ホールディング名判定カウント
w_p_count = 0
w_p_company_content = ""
w_p_holding_content = ""

# エクセルファイルの開始行
start_row = 5
count_row = start_row


# 以前に実行した出力ファイルを削除。出力ファイルが開いている場合は例外処理でメッセージを出力して終了する
def remove_output_file():
    is_file = os.path.isfile(output_file)
    if is_file:
        try:
            os.remove(output_file)
        except PermissionError as e:
             print(f"{output_file}が開いています。閉じてください")
             sys.exit()

# 出力ファイルのコピーメソッド
def output_file_copy():
    shutil.copy2(template_excel_file,output_file)

# タイトルのデータクレンジング関数
def title_cleansing(line):
    w_array1 = line.split(">")
    w_item = w_array1[1]
    w_item = w_item.replace('</td',"")
    return w_item

# 会社内容のデータクレンジング関数
def content_cleansing(line):
    w_array2 = line.split(">")
    w_content = w_array2[1]
    w_content = w_content.replace('<a href="',"")
    w_content = w_content.replace('tel:',"")
    w_content = w_content.replace('mailto:',"")
    w_content = w_content.replace('</td',"")
    w_content = w_content.replace('<br/',"")
    w_content = w_content.replace('"',"")
    return w_content

# pタグ内容のデータクレンジング関数
def p_tag_cleansing(line):
    w_array3 = line.split(">")
    w_p_content = w_array3[1]
    w_p_content = w_p_content.replace('</p',"")
    w_p_content = w_p_content.replace('・'," ")
    return w_p_content

# エクセルファイルへデータを書き込む関数
def output_data(w_item,company_content,holding_content):
    # 出力先エクセルファイルを開く。
    wb = op.load_workbook(output_file)
    sh_name = "会社概要"
    ws = wb[sh_name]
    ws.cell(row=count_row,column=2).value = w_item
    ws.cell(row=count_row,column=3).value = company_content
    ws.cell(row=count_row,column=4).value = holding_content
    wb.save(output_file)

# メイン処理
remove_output_file()
output_file_copy()

#タグ抽出ファイルを開く
file_data = open(input_file,"r",encoding="utf-8")
for line in file_data:
    line = line.replace("\n","")
    result1 = re.match(pattern1,line)
    result2 = re.match(patturn2,line)
    result3 = re.search(patturn3,line)
    result4 = re.search(patturn4,line)
    result5 = re.search(patturn5,line)
    # lineにrecord1のレコードが読み取られたら会社名からホールディングス名に変更する
    if line == record1:
        w_p_count += 1
    # 事業内容のtdタグを読み取ったらpタグの内容をエクセルファイルに出力する
    if line == record2:
        output_data(w_item,w_p_company_content,w_p_holding_content)
        count_row += 1
        w_p_count = 0
    # 項目データクレンジング関数を呼び出す
    if result1:
        w_count = 0
        w_item = title_cleansing(line)
    # 会社、ホールディングスの項目内容データクレンジング関数を呼び出す(pタグ以外)
    if (result2 and result3) or (result2 and result4):
        w_content = content_cleansing(line)
        if w_count == 0:
            w_company_content = w_content
        elif w_count == 1:
            w_holding_content = w_content
        w_count += 1
    # 会社、ホールディングスの項目内容データクレンジング関数を呼び出す(pタグ)
    if result5:
        w_p_content = p_tag_cleansing(line)
        if w_p_count == 1:
            w_p_company_content +=  w_p_content
        elif w_p_count == 2:
            w_p_holding_content += w_p_content


    if w_count == 2:
        output_data(w_item,w_company_content,w_holding_content)
        count_row += 1

file_data.close()
