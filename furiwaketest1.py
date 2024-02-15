# 人事異動作業担当者割り振りツール
#  新規作成  2024/2/14
#  Create by Accenture Satellight Namamugi


# ライブラリを読み込む
import openpyxl as op
from openpyxl import load_workbook
import re
import datetime
import shutil
import sys


# メンバー配列リスト
array1 = []
dict1 = {}

infile = r"C:\\nikkei\\furiwake\\担当振り分け.xlsx"
sh_name = "検索結果"
num = 0


# 担当振り分けエクセルファイルのデータ初期化
def inisial():
    start_row = 3
    max_row = start_row
    wb = op.load_workbook(filename = infile)
    ws = wb[sh_name]
    while True:
        work_num = ws.cell(row=max_row,column=9).value 
        if work_num == None:
            break
        else:
            ws.cell(row=max_row,column=2).value = None
            ws.cell(row=max_row,column=3).value = None
            ws.cell(row=max_row,column=4).value = None
            ws.cell(row=max_row,column=6).value = None
            ws.cell(row=max_row,column=7).value = None
            ws.cell(row=max_row,column=8).value = None
            ws.cell(row=max_row,column=9).value = None
            max_row += 1
    
    try:    
        wb.save(infile)
    except PermissionError:
            print(f"{infile}エクセルファイルを閉じてください") 
            sys.exit()
    
                   

# マクロテンプレートのデータを作業振り分けファイルにコピーする
def datacopy():
    start_row = 3
    max_row = start_row
    A_count = 0
    tempfile = r"C:\\nikkei\\furiwake\\担当振り分けテンプレート.xlsm"
    wb1 = op.load_workbook(filename = tempfile,read_only="yes")
    ws1 = wb1[sh_name]
    wb2 = op.load_workbook(filename = infile)
    ws2 = wb2[sh_name]
    while True:
        work_num = ws1.cell(row=max_row,column=9).value 
        if work_num == None:
            break
        else:
            company = ws1.cell(row=max_row,column=2).value
            url_value = ws1.cell(row=max_row,column=3).value
            ymd = ws1.cell(row=max_row,column=4).value
            url = ws1.cell(row=max_row,column=6).value
            company_name = ws1.cell(row=max_row,column=7).value
            flag = ws1.cell(row=max_row,column=8).value
            value1 = ws1.cell(row=max_row,column=9).value
            ws2.cell(row=max_row,column=2).value = company
            ws2.cell(row=max_row,column=3).value = url_value
            ws2.cell(row=max_row,column=4).value = ymd
            ws2.cell(row=max_row,column=6).value = url
            ws2.cell(row=max_row,column=6).hyperlink = url
            ws2.cell(row=max_row,column=7).value = company_name
            ws2.cell(row=max_row,column=8).value = flag
            ws2.cell(row=max_row,column=9).value = value1
            if value1 == "A":
                A_count += 1 
            max_row += 1
            

    #エクセルファイルを保存する。開いている場合は例外処理で終了する。 
    # ファイルが開いていたら、ファイル名を表示して処理を中止する。 
    try:    
        wb2.save(infile)
    except PermissionError:
        print(f"{infile}エクセルファイルを閉じてください") 
        sys.exit()
    return A_count    
    

# 日付付きの担当振り分けファイルを作成する。
def main(A_count):
    num = 0
    start_row = 3
    max_row = start_row
    work_num = 0
    work_data = 0
    count = 0
    j = 0
    dt = datetime.datetime.now()
    date1 = dt.strftime("%Y%m%d")
    outfile = r"C:\\nikkei\\furiwake\\担当振り分け" + date1 + ".xlsx"
    shutil.copy(infile,outfile)
    wb = op.load_workbook(filename = outfile)
    ws = wb[sh_name]
    A_workload = 0
    work_A_count = 0
    # 人数入力
    while True:
        try:
            num = int(input("作業人数を入力してください\n"))
            if num < 1:
                raise ValueError
            else:
                break
        except ValueError:
            print("1以上の正の整数を入力してください")    
        except TypeError:
            print("文字列を入力してください")
    
    #作業担当者を入力
    for i in range(0,num):
        print("作業担当者の名前を入力してください")
        while True:
            # 例外処理でスペースだけ、タブスペースだけ、何も入力しない場合は例外処理で入力されるまでループ
            try:
                data1 = input()
                data1 = data1.replace(" ","")
                data1 = data1.replace("\t","")
                if len(data1) == 0:
                    raise ValueError
                else:
                    array1.append(data1)
                    count += 1
                    break
                    
            except ValueError:
                print("作業者が入力されていません。")
          
    # 本日の検索結果のデータがNoneになるまで1行ずつ読み込む
    while True:
        work_num = ws.cell(row=max_row,column=9).value 
        if work_num == None:
            work_data += max_row - start_row 
            break
        else:
            max_row += 1
    #　Dictionaryに名前と初期化した値を代入する 
    for i in range(num):
        name = array1[i]
        dict1[name] = ""
    A_eval = int(A_count / num) 
    # データ開始行とデータ終了行を読み込み、作業量の重みが「A」データは1件または2件担当者をアサインし、
    # 作業量の軽い転記作業にアサインしない。
    #  作業量の重みが「B」、「C」の時は複数件の転記作業をアサインする
    for data in range(start_row,max_row):
        if j == count:
            j = 0

        while True:
            name1 = array1[j]
            workload = ws.cell(row=data,column=9).value
            value1 = dict1[name1]
            value1_count = len(value1)
            # Aの件数を人数で割った商が0なら作業量「A」を1件割り振ったメンバーは「B」、「C」
            # は割り振らない
            # Aの件数を人数で割った商が1なら作業量「A」を21件割り振ったメンバーは「B」、「C」
            # は割り振らない
            
            if A_eval == 0:
                A_workload = 1
            elif A_eval == 1:
                A_workload = 2
            if value1_count > 0:
                work_A_count = 0 
                
                for k in range(0,value1_count):
                    w_str1 = value1[k]
                    if w_str1 == "A":
                        work_A_count += 1

            if work_A_count == A_workload:
                j += 1
            else:
                break    
        value1 += workload
        ws.cell(row=data,column=10).value = name1    
        dict1[name1] = value1
        print(dict1)
        j += 1
    #エクセルファイルを保存する関数を呼び出す。 
    try:    
        wb.save(outfile)
    except PermissionError:
        print(f"{outfile}エクセルファイルを閉じてください") 
        sys.exit()

# メイン関数
if __name__ == '__main__':
    inisial()
    numA = datacopy()
    main(numA)
    
