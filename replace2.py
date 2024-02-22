###############################################################
####     ログイン数のログファイルを「ログイン数」_org.xlsmファイルに書き込む    #######
####     Create   2024/2/22
####     Author   takao.hattori
###############################################################
# モジュールのインポート
import re
import os
import openpyxl

# 変数の定義
out_file = "ログイン数_org.xlsm"
wb = openpyxl.load_workbook(out_file,keep_vba=True)
ws = wb['ログイン数']
start_row = 2

# ログイン数ログファイル名取得関数
def search_logfile():
    DIR_PATH = os.getcwd()
    file_list = os.listdir(DIR_PATH)
    search_file = ""
    for i in list(file_list):
        result1 = re.search('log',i)
        if result1:
            search_file = i
    return search_file        

# 前月の「ログイン」シートに記載されているログイン数データ数取得関数
def get_previous_row():
    
    count_row = start_row
    while True:
        if ws.cell(row=count_row,column=2).value == None:
            break
        else:
            count_row += 1
            
    max_row = count_row
    return max_row    

# 前回のデータを削除する
def clear_data(max_row):
    for i in range(start_row,max_row):
        ws.cell(row=i,column=2).value = None
        ws.cell(row=i,column=3).value = None
        ws.cell(row=i,column=4).value = None
    wb.save(out_file)

#　今回作業分のログインユーザー数をExcelに書き込む関数。
def insert_data(input_file):
    with open(input_file,mode="r") as f:
        j = start_row
        for line in f:
            line1 = line.replace("\n","")
            print(line1)
            array1 = line1.split(' ')
            date1 = array1[0]
            hour1 = array1[1]
            login_count = int(array1[2])
            ws.cell(row=j,column=2).value = date1
            ws.cell(row=j,column=3).value = hour1
            ws.cell(row=j,column=4).value = login_count
            j += 1
        wb.save(out_file)    

 
if __name__ == '__main__':
    in_file = search_logfile()
    max_row = get_previous_row()
    clear_data(max_row)    
    insert_data(in_file)
