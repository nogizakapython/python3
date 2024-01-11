####################################################
###   お題抽選ツール
###
####################################################

#ランダムモジュールを呼び出す
import random
# エクセルへの書き込みモジュールを呼び出す
import openpyxl
# 関連付けられたアプリケーションでファイルを開くモジュールを呼び出す
import subprocess
# 非同期処理用モジュールを呼び出す
import asyncio
# タイムモジュールを呼び出す
import time
# システムモジュールを呼び出す
import sys

#配列を定義する
array1 = ["大園玲","小坂菜緒","正源司陽子","守屋麗奈","佐々木久美","山下瞳月","佐々木美玲","梅澤美波","筒井あやめ","遠藤さくら","中西アルノ","村井優","井上和","丹生明星","小林由依","山﨑天"]

array2 = ["好きな小説家","好きなお笑い芸人","好きなファッションブランド","好きなコスメブランド","好きな赤文字系ファッション雑誌","好きなロックバンド","住みたい街","好きなランニングシューズ","好きなおにぎりの具","納豆餅に使う納豆は小粒派、ひきわり派","PCはWindows派、mac派、Linux派?","一度は行きたい国","好きなゲーム","手軽にできる料理","初心者にお薦めのプログラミング言語","初詣でおみくじを引く?"]

#エクセルファイルのファイル名変数
out_file = "odai.xlsx"
wb = openpyxl.load_workbook(out_file)
#書き込むワークシートを定義する
ws = wb['select']
#行数の定義
start_row = 3
# 待ち時間変数
sleep_time = 10
#演出待ち時間変数
play_time = 2

# 非同期処理で1秒後に「抽選結果の発表です」を出力し、９秒待つ
async def start_msg():

  print("抽選結果の発表です")
  div1()
  await asyncio.sleep(sleep_time)

# 非同期処理で10秒間演出する
async def message_out():
    for k in range(0,5):
      await asyncio.sleep(play_time)
      print("ﾄﾞﾗ" * 20)
      div1()


async def main():

  task1 = asyncio.create_task(
        start_msg()
  )

  task2 = asyncio.create_task(
        message_out()
  )

  print(f"started at {time.strftime('%X')}")
  await task1
  await task2
  print(f"ended at {time.strftime('%X')}")

#配列の数の比較関数
def array_count(num1,num2):
    if num1 == num2:
      return "OK"
    else:
      return "NG"

def div1():
  print("=" * 40)


ans =  array_count(len(array1),len(array2))
if ans == "NG":
  print("人数とお題のタイトル数が違うので、array1とarray2の数を合わせてください")
  sys.exit()

N = len(array1)

#配列の要素をランダムに並べる
random_array1 = random.sample(list(set(array1)),N)
random_array2 = random.sample(list(set(array2)),N)


#ランダムに並べた名前を1行ずつエクセルファイルに出力する。
for k in range(0,N,1):
  man1 = random_array1[k]
  title1 = random_array2[k]
  try:
    with open(out_file,mode='a') as f:
      #エクセルファイルに追記モードで書き込む
      ws.cell(row=start_row,column=2).value = man1
      ws.cell(row=start_row,column=3).value = title1
      start_row += 1
  except:
    print("エクセルファイルが開いたままです。閉じてください")
    exit()

#エクセルファイルを保存する
wb.save(out_file)

# 非同期処理で
asyncio.run(main())


#エクセルファイルを開く
subprocess.Popen(['start',r'odai.xlsx'],shell=True)
