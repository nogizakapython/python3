####################################################
###   お題抽選ツール
####################################################

#ランダムモジュールを呼び出す
import random
import openpyxl
import subprocess

#配列を定義する
array1 = []
array2 = []
#人数変数
N = 0
#エクセルファイルのファイル名変数
out_file = "prezen1.xlsx"
wb = openpyxl.load_workbook(out_file)
#書き込むワークシートを定義する
ws = wb['select']
#行数の定義
start_row = 3


#正の整数が入力されるまで繰り返す
while True:
  try:
    N = int(input("登録人数を指定してください\n"))
    if N <= 0:
      raise ValueError("error!")
    else:
      break
  except ValueError as e:
    print("文字列、小数、負の整数、0を入力しないでください")
    print("正の整数を入力してください")

#配列に名前を登録する
for i in range(0,N):
  name = input("名前を入力してください\n")
  array1.append(name)

#配列にお題を登録する
for j in range(0,N):
  title = input("プレゼンのお題を入力してください\n")
  array2.append(title)

#配列の要素をランダムに並べる
random_array1 = random.sample(list(set(array1)),N)
random_array2 = random.sample(list(set(array2)),N)

#ランダムに並べた名前を1行ずつエクセルファイルに出力する。
for k in range(0,N,1):
  man1 = random_array1[k]
  title1 = random_array2[k]
  try:
    with open(out_file,mode='a') as f:
      #エクセルファイルに追記モードで書き込む
      ws.cell(row=start_row,column=2).value = man1
      ws.cell(row=start_row,column=3).value = title1
      start_row += 1
  except:
    print("エクセルファイルが開いたままです。閉じてください")
    exit()
#エクセルファイルを保存する
wb.save(out_file)
#エクセルファイルを開く
subprocess.Popen(['start',r'prezen1.xlsx'],shell=True)
