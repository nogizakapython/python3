try:
    import datetime
    import os
    #Excel
    import  openpyxl
    #ウィンドウメッセージ
    import tkinter as tk
    from tkinter import messagebox
    #クリップボード値の操作
    import pyperclip
    # 時間を計るライブラリをインポート
    import time
    # WebDriverライブラリをインポート
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.edge.service import Service
    from webdriver_manager.microsoft import EdgeChromiumDriverManager # type: ignore

    today =datetime.datetime.now()
    tymd = today.strftime('%Y-%m-%d %H:%M:%S')

    ##この実行ファイルの存在する位置（パス）を確認し、そこから必要ファイルの格納先を取得する
    masterPath = os.path.abspath(__file__)
    masterPath = masterPath[2:]
    ind = masterPath.index("RPA-ACNChat_for_Selection_of_new_graduates")
    masterPath = masterPath[:ind]
    masterPath = "C:"+ masterPath
    xlmasterPath = masterPath + 'Book1.xlsx'
    recXt = masterPath+'実行履歴.txt'


    with open(recXt,'a') as t: 
        t.write('\n'+tymd+': ')

    wb = openpyxl.load_workbook(xlmasterPath)
    ws = wb["Sheet1"]
    maxRow = ws.max_row
    print(maxRow)
    i =2

    '''#チェックボックスを確認
    check_box = driver.find_element(By.ID, "checkbox-0")
    # 指定した要素（チェックボックス）がチェックされてない場合はチェック
    if not check_box.is_selected():
        # 非選択の場合はJavaScriptでクリックする
        driver.execute_script("arguments[0].click();", check_box)
    '''


    while i <= maxRow:
            
            # Edgeを指定する
            if i == 2:
                driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
                #driver = webdriver.Edge()
                # Edgeを開いてAccenture Chatにアクセスする
                driver.get('https://peerworker.ai')
                #待機時間設定最大200秒
                driver.implicitly_wait(200)
                #SIGNINボタンを確認
                SignIN_button = driver.find_element(By.XPATH, '//*[@id="sign"]')
                #クリック
                SignIN_button.click()
                driver.get('https://peerworker.ai/apps/051a3244-e488-4f8e-abae-c08ef6a2c8e3/run?appVersion=1')

            #Excelからキーワードを取得
            kWord = ws.cell(row=i, column=3)
            kWord = kWord.value
            print(kWord)
            
            #プロンプト
            pText = f"""以下内容を踏まえて、使用するケース問題を#出力形式にしたがって作成してください。

            <作成条件>
            ・テーマは、#Caseのテーマについて
            ・アクセンチュアらしさを体現した、時代を反映したデジタル関連のお題であること
            ・アイデアが収束できるよう、ある程度定量的な視点をもって考えることができるCaseであること
            ・フェルミ推定×施策立案を組み合わせたお題であること
            ※注意事項※
            ・回答するのに特定の知識を要さないもの
            ・居住地や性別等の属性による知識格差が発生しすぎないもの（例：特定の国の政策・特定の市の状況）
            ・特定のクライアントが想起されるお題とすること（但し、企業名・サービス名は伏せること）
            ・解のあるお題であること
            ・宗教・出生地・ジェンダー関連のお題など、センシティブな内容を文内に含まないこと　
            ＜補足＞
            ・日常生活や時事問題に関連したテーマの方が、学生にとっても公平な選考となり、面接評価に十分なアウトプットとなる傾向があります

            #出力形式
            背景・課題・質問を合計3文程度で、Caseの文章のみ出力
            （背景・課題・質問の言葉は不要）

            #Caseのテーマ
            {kWord}
            """
            strs = pText.split('\n')

            print("test")

            #入力箇所を確認
            if i != 2:
                # 5秒待機する
                time.sleep(5)
            WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')))
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            print('ttest')
            # 5秒待機する
            #time.sleep(3)
            #検索ボックスへ入力
            #inputBox.clear()
            inputBox.send_keys(Keys.ARROW_LEFT)
            inputBox.send_keys(Keys.ARROW_LEFT)
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            inputBox.send_keys(Keys.BACK_SPACE)
            inputBox = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div[22]')
            '''
            for str in strs:
                inputBox.send_keys(str)
                inputBox.send_keys(Keys.SHIFT,Keys.ENTER)
            '''
            inputBox.send_keys(kWord)

            #実行ボタンを確認
            execute_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[3]/div/div[1]/div')
            #クリック
            execute_button.click()
            

            '''
            #三点ボタンを確認
            dot_button = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div/div[2]/main/div[1]/div/div/div[1]/div/div[2]/div/button/span[3]/i')
            #クリック
            dot_button.click()
            '''


            #コピーボタンを確認
            try:
                WebDriverWait(driver, 300).until(EC.text_to_be_present_in_element((By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[2]/div[3]/div/div[1]/div'),'一括実行する'))
                copy_button = driver.find_element(By.CSS_SELECTOR, '.base-button.center.with-opacity.icon-button.button.copy-button')
                #WebDriverWait(driver, 300).until(EC.element_to_be_clickable(By.CSS_SELECTOR, '.base-button.center.with-opacity.icon-button.button.copy-button'))
                #driver.execute_script("arguments[0].click();", copy_button)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                copy_button.click()
                wCheck = pyperclip.paste()

                while wCheck == 0:
                    # 秒待機する
                    time.sleep(3)
                    copy_button.click()
                    wCheck = pyperclip.paste()

                #結果をエクセルへ貼り付け
                rCel = ws.cell(row=i, column=4)
                rCel.value = pyperclip.paste()

                wb.save(xlmasterPath)
                pyperclip.copy(0)


                #画面のリロード
                #reload_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[2]/div/div/div/div[3]/div')
                #driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                #reload_button.click()
                #driver.execute_script("arguments[0].click();", reload_button)
                #driver.quit()
                driver.refresh()
            finally:
                tst = 1

            


            i += 1
            # 秒待機する
            time.sleep(3)

    # 画面を閉じる
    driver.quit()

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo('実行終了',  '作業が終了しました。')
    
    with open(recXt,'a') as t: 
        t.write('完了')

except:
    import traceback
    e = traceback.format_exc() 
    e=str(e)

    with open(recXt,'a') as t: 
        t.write(f'[Error]{e}')