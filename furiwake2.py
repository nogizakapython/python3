import math
import openpyxl as op
from openpyxl import load_workbook

dic1 = {"小坂菜緒":0,"富田鈴花":0,"山口陽世":0,"齊藤京子":0}
w_name = []

wb = op.load_workbook("sample1.xlsx")
sh_name1 = 'Sheet1'
ws = wb[sh_name1]


num = len(dic1)
max_row = 0
start_row = 3
work_data = 0
work_sum = 0
j = 0

max_row = start_row

while True:
        work_num = ws.cell(row=max_row,column=3).value
        if work_num == None:
            work_data += max_row - start_row 
            break
        else:
            work_num = int(work_num)

            work_sum = work_sum + work_num
            max_row += 1
       

avg = math.ceil(work_sum / num)

for name in dic1.keys():
     w_name.append(name)

count = len(w_name)

for data in range(start_row,max_row):
    if j == count:
         j = 0

    name1 = w_name[j]
    #print(name1)
    price = ws.cell(row=data,column=3).value
    ws.cell(row=data,column=4).value = name1
    j += 1
    #print(j)
    

wb.save("sample1.xlsx")
     



