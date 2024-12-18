# 行削除テスト

# ライブラリのインポート
import openpyxl

# Global変数
# ファイル名
file_name = "data1.xlsx"
# ExcelのWorkbookインスタンスの呼び出し
wb = openpyxl.load_workbook(file_name)
# ExcelのWorksheet変数を定義する
ws = wb['Sheet1']

# 行の最終行を取得する関数
def end_row():
    # 開始行変数
    start_num = 4
    # カウント変数
    i = start_num
    while True:
        dte_str = ws.cell(i,4).value
        if dte_str != None:
            i += 1
        else:
            break
    return i-1        

finish_row_num = end_row()

count = 4
while True:
    div = ws.cell(count,4).value
    # 部署が「Tecnplogy」の時、１行削除する
    if div == "Tecnology":
        # print("削除するよ～")
        # print(count)
        ws.delete_rows(count)
    elif div == None:
        break    
    else:    
        count += 1

# 削除処理後、ファイルを保存する
wb.save(file_name)    