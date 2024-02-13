# 担当割り振りツール
#  新規作成  2024/2/13


# ライブラリを読み込む
import openpyxl as px
from openpyxl import load_workbook



# メンバー配列リスト(作業者6人まで)
array1 = {}
array2 = []

infile = r"C:\\nikkei\\振り分け\\担当振り分けテンプレート.xlsx"
sh_name = "検索結果"
wb = load_workbook(filename = infile)
ws = wb[sh_name]

def main():
    num = 0
    work_data = 0
    work_sum = 0
    start_row = 3
    max_row = start_row
    count = 0
    while True:
        try:
            num = int(input("作業人数を入力してください\n"))
            if num < 1:
                raise ValueError
            else:
                break
        except ValueError:
            print("1以上の正の整数を入力してください")
        except TypeError:
            print("文字列を入力してください")
    for i in range(0,num):
        print("作業担当者の名前を入力してください")
        while True:

            try:
                data1 = input()
                if len(data1) == 0:
                    raise ValueError
                else:
                    new_key = str(data1)
                    array1[new_key] = 0
                    count += 1
                    break

            except ValueError:
                print("作業者が入力されていません。")


    while True:
        work_num = ws.cell(row=max_row,column=2).value
        if work_num == None:
            work_data += max_row - start_row
            break
        else:
            work_num = int(work_num)
            array2.append(work_num)
            work_sum = work_sum + work_num
            max_row += 1







if __name__ == '__main__':
    main()


