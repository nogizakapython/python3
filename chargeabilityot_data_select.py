##########################################################
#  chargeability OT Operations以外のデータ削除処理       ###
#  新規作成 2024/12/18  takao.hattori                   ###
##########################################################

# ライブラリのインポート
import openpyxl
import os
import codecs
import re
import shutil
import time

# Global変数
# パス変数の定義
file_path = 'C:\chargeability'
# 一覧ファイル
file_name = 'file_list.txt'

def make_file_list_array():
    # ファイル名配列
    work_array = os.listdir(file_path)
    return work_array

# 前回のファイル一覧ファイル削除関数
def before_remove_file_list():
    # 前回のファイルリストを削除
    os.remove(file_name)

# ファイルリストの作成と、FlashまたはFinishのOTデータファイル名取得関数
def get_base_otfile(file_list_array):    
    # ファイル名返り値変数
    return_file_name = ""
    
    # ファイル一覧作成
    for filename in file_list_array:
        print(filename,file=codecs.open(file_name,'a','utf-8'))

    # ファイル一覧から、「FlashまたがFinishで始まるファイルをヒットさせる。
    with open(file_name,encoding="utf-8") as f:
        for line1 in f:
            line1 = line1.replace("\n","")
            result1 = re.match(r'作業用*',line1)
            # print(result1)
            if result1:
                return_file_name = line1
                break
    return return_file_name     

# 作業用ファイル作成関数 
def make_work_target_file(base_target_file,change_target_file):
    # 作業用のファイルがあれば削除する
    if os.path.isfile(change_target_file):
        os.remove(change_target_file)
    # 原本のOTファイルをコピーし、作業用ファイルを作成する。
    shutil.copy(base_target_file,change_target_file)


def main():
    # ファイル一覧配列作成
    file_list_array = make_file_list_array()
    # 前回のファイル一覧ファイルがあれば削除
    before_remove_file_list()
    # 原本のFlashまたはFinishのOTファイル名取得関数
    change_target_file = get_base_otfile(file_list_array)
    print(change_target_file)
    # change_target_file = "work_" + base_target_file
    # 作業用OTファイル作成
    # make_work_target_file(base_target_file,change_target_file)

    # print(change_target_file)
    # Workbookインスタンス変数
    wb = openpyxl.load_workbook(change_target_file)
    # ExcelのWorksheet変数を定義する
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 開始行をセット
    count = 3
    procedure_count = 3
    while True:
        div = ws.cell(count,7).value
        # 部署が「Tecnplogy」の時、１行削除する
        if div == None:
            break
        elif div != "Operations" and div != "Operations-FT" and div != "Operations-RS":
            ws.delete_rows(count)
        else:    
            count += 1

        procedure_count += 1
        if procedure_count % 10 == 0:
            wb.save(change_target_file)    

    # 削除処理後、ファイルを保存する
    wb.save(change_target_file)
    print("削除処理完了だよ") 


if __name__ == '__main__':
    main()    