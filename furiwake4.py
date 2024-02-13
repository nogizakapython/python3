# タスク振り分けツール
#  新規作成  2024/2/13


# ライブラリを読み込む
import openpyxl as op
from openpyxl import load_workbook
import re


# メンバー配列リスト
array1 = []
dict1 = {}

infile = r"C:\\dcareer\\python\\タスク割り当て.xlsx"
sh_name = "タスク"
wb = op.load_workbook(filename = infile)
ws = wb[sh_name]

def main():
    num = 0
    start_row = 3
    max_row = start_row
    work_num = 0
    work_data = 0
    count = 0
    j = 0
    while True:
        try:
            num = int(input("作業人数を入力してください\n"))
            if num < 1:
                raise ValueError
            else:
                break
        except ValueError:
            print("1以上の正の整数を入力してください")
        except TypeError:
            print("文字列を入力してください")
    for i in range(0,num):
        print("作業担当者の名前を入力してください")
        while True:

            try:
                data1 = input()
                if len(data1) == 0:
                    raise ValueError
                else:
                    array1.append(data1)
                    count += 1
                    break

            except ValueError:
                print("作業者が入力されていません。")

    # タスク数
    while True:
        work_num = ws.cell(row=max_row,column=3).value
        if work_num == None:
            work_data += max_row - start_row
            break
        else:
            max_row += 1
    # Dictionaryに名前と初期化した値を代入する
    for i in range(num):
        name = array1[i]
        dict1[name] = ""
    print(count)
    # エクセルのデータ件数分読み込む
    for data in range(start_row,max_row):
        if j == count:
            j = 0
        print(j)
        # 難易度Aのタスクを1件だけもつようにタスクを割り振る
        while True:
            name1 = array1[j]
            workload = ws.cell(row=data,column=3).value
            value1 = dict1[name1]
            result1 = re.search("A",value1)
            if result1:
                j += 1
            else:
                break

        value1 += workload
        # エクセルファイルに書き込む
        ws.cell(row=data,column=4).value = name1
        # 連想配列の値を更新する
        dict1[name1] = value1

        j += 1
    wb.save(infile)

# メイン関数
if __name__ == '__main__':
    main()


