#######   Freestyle 企業HPから人事情報を取得する　###########
#######   新規作成  2024/2/11  ##########
#######   Author  乃木坂好きのITエンジニア ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
# Excelファイルへ書き込むライブラリをインポート
import openpyxl as op
import codecs
from time import sleep
import sys
import shutil


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')

# ソースコード出力ファイル
file_name = "freestyle" + date1 + ".txt"
out_file = "freestyle.txt"

# スクレイピング対象URL
target_url = 'https://freestyles.jp/news/'
# エクセルファイル書き出し開始行変数の定義
max_row = 5
row_count = 0
# テンプレートエクセルファイルの定義
import_file = "フリースタイル_yyyymmdd.xlsx"
# 出力先エクセルファイル
export_file = "フリースタイル_" + date3 + ".xlsx"


# Chromeを指定する
driver = webdriver.Chrome()

try:
    shutil.copy2(import_file,export_file)
except FileNotFoundError:
    print("リネーム前のファイルが存在しません")
    sys.exit()
except PermissionError as e:
    fname = export_file
    print("テンプレートファイルが開いています。閉じてから実行してください")
    sys.exit()

# Chromeを開いて企業HPのプレスリリースにアクセスする
try:
    driver.get(target_url)
    sleep(5)

    for i in range(1,16):

        xpath_str1 = '/html/body/main/div[3]/div/ul[2]/li[' + str(i) + ']/a/div[2]/div[1]/div[1]'

        try:
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        except:
            break
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        xpath_str2 = '/html/body/main/div[3]/div/ul[2]/li[' + str(i) + ']/a/div[2]/div[2]'
        element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
        print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))


except EnvironmentError as e:
    str100 = e


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
# 以前のスクレイピング結果テキストファイルが残っていれば削除
if file_exist:
   os.remove(out_file)

shutil.copy2(file_name,out_file)


# スクレイピングしたファイルをunicodeで開く
fileobj = open(out_file,encoding="utf-8")
# スクレイピングファイルを1行ずつEOFまで読み込む
while True:
    line1 = fileobj.readline()
    # 改行コードを除去
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break
    # 抽出条件
    result1 = re.match('<div class="time">',line1)
    result2 = re.match('            ',line1)
    # 年月日データのデータをクレンジングする。
    if result1:
        w_array1 = line1.split(">")
        w_y = w_array1[1]
        w_y = w_y.replace("</div","")
        w_y = w_y.replace(".","/")
        w_ymd = w_y
        #print(w_ymd)
    # タイトルを抽出する
    if result2:
        w_array2 = line1.split("<")
        w_title = w_array2[0]
        w_title = w_title.replace(" ","")
        # エクセルファイルに書き出す
        wb = op.load_workbook(export_file)
        sh_name = 'freestyle'
        ws = wb[sh_name]
        ws.cell(row=max_row,column=2).value = w_title
        ws.cell(row=max_row,column=3).value = w_ymd
        # 出力するエクセルの行を1行足す
        max_row += 1
        # エクセルファイルの保存
        try:
            wb.save(export_file)
        except PermissionError as e:
            fname = export_file
            print("書き込み先ファイルが開いています。ファイルを閉じてください")
            sys.exit()
