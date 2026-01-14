#######   Kubota 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/26  ##########
#######   修正      2024/06/19 空白のNone行対応 #############  
#######   Author  takao.hattori ###########


# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


out_file = "kubota.txt"
date_str = ""
w_title = ""

web_url = 'https://pps-net.org/statistics/crude-oil'
target_url = ""
max_row = 5
base_file = ""
export_file = "石油価格値段.xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
start_row_num = 1
col_num = 2
row_num = start_row_num


# Chromeを指定する
driver = webdriver.Chrome()

target_url = web_url
try:
    driver.get(target_url)
    sleep(3)


    try:
        xpath_str1 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[1]/th[2]'
        xpath_str2 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[2]/td[1]'
        xpath_str3 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[3]/td[1]'
        xpath_str4 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[4]/td[1]'
        xpath_str5 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[5]/td[1]' 
            
                     
    except:
         print(1)
        
    element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
    element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
    element_str3 = driver.find_element(by=By.XPATH,value=xpath_str3)
    element_str4 = driver.find_element(by=By.XPATH,value=xpath_str4)
    element_str5 = driver.find_element(by=By.XPATH,value=xpath_str5)
    
    
    ymd = element_str1.get_attribute("outerHTML")
    ymd = ymd.replace("<th>","")
    ymd = ymd.replace("</th>","")
    w_array = ymd.split('/')
    month = w_array[1]
    work_month = ""
    
    if month == "1":
        work_year = w_array[0][-2:]
        work_month = work_year + "年" + month + "月"
    else:
        work_month = month + "月"

    wb = op.load_workbook(export_file)
    sh_name = 'Sheet1'
    ws = wb[sh_name]
    ws.cell(row=row_num,column=col_num).value = work_month

    row_num += 1
    dubai_accout = element_str2.get_attribute("outerHTML")
    dubai_accout = dubai_accout.replace(' $/バレル',"")
    dubai_accout = dubai_accout.replace('<td>',"")
    dubai_accout = dubai_accout.replace('</td>',"")

    ws.cell(row=row_num,column=col_num).value = float(dubai_accout)

    row_num += 1
    blend_accout = element_str3.get_attribute("outerHTML")
    blend_accout = blend_accout.replace(' $/バレル',"")
    blend_accout = blend_accout.replace('<td>',"")
    blend_accout = blend_accout.replace('</td>',"")

    ws.cell(row=row_num,column=col_num).value = float(blend_accout)

    row_num += 1
    wti_accout = element_str4.get_attribute("outerHTML")
    wti_accout = wti_accout.replace(' $/バレル',"")
    wti_accout = wti_accout.replace('<td>',"")
    wti_accout = wti_accout.replace('</td>',"")

    ws.cell(row=row_num,column=col_num).value = float(wti_accout)

    row_num += 1
    opec_accout = element_str5.get_attribute("outerHTML")
    opec_accout = opec_accout.replace(' $/バレル',"")
    opec_accout = opec_accout.replace('<td>',"")
    opec_accout = opec_accout.replace('</td>',"")

    ws.cell(row=row_num,column=col_num).value = float(opec_accout)
    
    wb.save(export_file)   
except EnvironmentError as e:
    str100 = e     
     